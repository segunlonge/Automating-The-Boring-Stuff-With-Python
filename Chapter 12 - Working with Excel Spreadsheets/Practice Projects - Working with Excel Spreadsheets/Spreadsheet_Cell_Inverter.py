"""
Write a program to invert the row and column of the cells in the spreadsheet.
For example, the value at row 5, column 3 will be at row 3, column 5 (and vice versa).
This should be done for all cells in the spreadsheet.
"""

import openpyxl, os, math

#Set the Current Working Directory
os.chdir('C:\\Python Files\\Automating The Boring Stuff With Python\\Chapter 12 - Working with Excel Spreadsheets\\Practice Projects - Working with Excel Spreadsheets')

#Open Workbook and assign it to memory via object type
wb = openpyxl.load_workbook("Spreadsheet_Cell_Inverter.xlsx")
#Assign first sheet to sheet object with variable name 'sheet'
sheet = wb.worksheets[0]
#sheet = wb['Sheet1']

#open a new work book
wb_copy = openpyxl.Workbook()
#Assign first sheet to sheet object with variable name 'sheet'
sheet_copy = wb_copy['Sheet']

#Create a look up dictionary for looking up numberic to alphabetic for the purpose of getting a cell range reference
d = {1:"A",2:"B", 3:"C", 4:"D", 5:"E",6:"F", 7:"G", 8:"H", 9:"I",10:"J", 11:"K", 12:"L", 13:"M",14:"N", 15:"C",
     16:"O", 17:"P",18:"Q", 19:"R", 20:"S", 21:"T",22:"U", 23:"V", 24:"W", 25:"Y", 26:"Z"}

'''A function that takes in an integer n, looks up that integer in the dictionary value 'd' and returns a
letter to correspond to an Excel column which is used to create a range index e.g. A1

'''
def alpha_lookup(n):
    a = math.floor(n/26)
    b = n%26

    #If a is 0 then n is less than 26 and so return direct value
    if a == 0:
        return d[n]
    #This returns the the values of the first and second letter if input value greater than 26
    return d[a]+d[b]

#Get the maximum row and column number with values
maxcol = len(sheet[1])
print(maxcol)
maxrow = len(sheet['A'])
print(maxrow)

'''The loop uses the enumerate function to step through each row and each cell in each row to get
a numerical value that is; for each i (column) the loop goes through each cell (j).

Then the column(i) and the cell(j) are concatenated to make an Excel cell address range

The sheet.iter_rows method works on the sheet object to iterate through the cells of each
row e.g. A1 and B1 which gets translated to A1 and A2

'''
for i, row in enumerate(sheet.iter_rows(min_row=1, max_col=maxcol, max_row=maxrow)):
    for j, cell in enumerate(row):
        print(cell.value)
        index = alpha_lookup(i+1)+str(j+1)
        print(index)
        sheet_copy[index] = cell.value

#Save the workbook object called wb_copy
wb_copy.save("Spreadsheet_Cell_Inverter_Copy.xlsx")
        

