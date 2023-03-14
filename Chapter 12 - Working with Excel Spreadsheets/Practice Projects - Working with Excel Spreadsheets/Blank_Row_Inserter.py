"""
Create a program blankRowInserter.py that takes two integers and a filename string
as command line arguments. Let's call the first integer N and the second integer M.
Starting at row N, the program should insert M blank rows into the spreadsheet.

You can write this program by reading in the contents of the spreadsheet. Then, when
writing out the new spreadsheet, use a for loop to copy the first N lines.
For the remaining lines, add M to the row number in the output spreadsheet.
"""

# The shebang line which tells the computer to execute code using Python
#! python3

import sys, openpyxl, os

#Set the Current Working Directory
os.chdir('C:\\Python Files\\Automating The Boring Stuff With Python\\Chapter 12 - Working with Excel Spreadsheets\\Practice Projects - Working with Excel Spreadsheets')

#Check to see if all the command line arguments have been entered. If not exit
if len(sys.argv) < 4:
    print('Please enter all arguments')
    sys.exit()
else:
    print("Number of arguments:", len(sys.argv), "arguments")
    n_integer = int(sys.argv[2])
    m_integer = int(sys.argv[3])
    workbook_name = sys.argv[1]

wb = openpyxl.load_workbook('BlankRowInserter_Fruits.xlsx')
sheet = wb.worksheets[0]
#print(type(wb))

max_row = len(sheet[1])
max_col = len(sheet['A'])

from openpyxl.utils import get_column_letter, column_index_from_string
region_edge = get_column_letter(len(sheet[1]))+str(len(sheet['A']))

"""
setup a dynamic range to read in the contents of the spreadsheet. Use len(ws['A']) to find the last row and column
number.
"""

#Create an empty list to store all rows and then append each row list into the sheet_data list

sheet_data = []

for rowOfObjects in sheet['A1':region_edge]:
    row_data = []
    for cellObj in rowOfObjects:
        row_data.append(cellObj.value)
    sheet_data.append(row_data)

num_of_rows = len(sheet_data)

#open a new work book
new_wb = openpyxl.Workbook()
#set the first sheet as the working sheet
new_sheet = new_wb.worksheets[0]

#Loop through the first n items in the list and write to new workbook
#Use the slicing method for lists to take first n items

a = 0
b = 1

for i in sheet_data[:n_integer]:
    a = a + 1 
    for j in i:
        #Code to write into new workbook
        #print(j)
        new_sheet.cell(row=a, column=b).value = j
        b = b + 1
    b = 1
        
#Loop through the last n items in the list and write to new workbook
#Use the slicing method for lists to take last n items

a = 0
b = 1
target_row = n_integer + m_integer + 1

for l in sheet_data[n_integer:]:
    for m in l:
        #But shift the row down by n_integer + m_integer
        #print(m)
        new_sheet.cell(row=target_row, column=b).value = m
        b = b + 1
    b = 1
    target_row = target_row + 1

# Save workbook
new_wb.save('BlankRowInserter_Fruits_New.xlsx')
        
    
        






    
