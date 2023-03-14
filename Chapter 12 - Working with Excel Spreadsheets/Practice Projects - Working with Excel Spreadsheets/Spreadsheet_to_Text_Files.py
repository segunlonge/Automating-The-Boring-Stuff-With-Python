"""
Write a program that performs the tasks of the previous program in reverse order: The program
should open a spreadsheet and write the cells of column A into one text file, the cells of column B into
another text file, and so on
"""

# The shebang line tells the computer to execute code using Python
#! python3

import os, openpyxl, fnmatch
os.chdir('C:\\Python Files\\Automating The Boring Stuff With Python\\Chapter 12 - \
Working with Excel Spreadsheets\\Practice Projects - Working with Excel Spreadsheets')

#open the required Excel workbook called 'Text_Files_to_Spreadsheet.xlsx'
wb = openpyxl.load_workbook('Text_Files_to_Spreadsheet.xlsx')
desired_sheet = wb.worksheets[0]

from openpyxl.utils import get_column_letter, column_index_from_string #Allows conversion fron numbers to letters
region_edge = get_column_letter(len(desired_sheet[1]))+str(len(desired_sheet['A']))

region_edge_2 = get_column_letter(len(desired_sheet[1]))+'1'

##print(region_edge)
##print(region_edge_2)


Key_List = []
Value_List = []
Header_List = []

region_data = {}

#Get cell coordiates and cell values from the Excel region and populate in 2 lists for later dictionary population
for rowOfCellObjects in desired_sheet['A1':region_edge]:
      for cellObj in rowOfCellObjects:
          #print(cellObj.coordinate, cellObj.value)
          Key_List.append(cellObj.coordinate)
          Value_List.append(cellObj.value)

#Get the cell cordinates and strip out the number part
for rowOfCellObjects in desired_sheet['A1':region_edge_2]:
      for cellObj in rowOfCellObjects:
            Header_List.append(cellObj.coordinate[:1])

#print(Header_List)
            
##print(Key_List)
##print(Value_List)

#Assign each cell reference as a key with the cell contents as the value
for x in range(len(Key_List)):
      region_data[Key_List[x]] = Value_List[x]

print(region_data)

# Get all the values per column as per Excel file using the cell references and wildcard matching before writing to text files
for q in Header_List:
      for k, v in region_data.items():
            if q+'1' == k:    #Take the first row items and use that as the text file name
                  TextFile = open(v+'.txt','w')
            '''[:1] gets the first character of the key value and appends an * to it for the wildcard search in fnmatch
            where the values are 'None' replace with blank space and only take values which are not in the first row
            '''
            if fnmatch.fnmatch(q, k[:1]+'*') and str(v) == 'None' and q+'1' != k:
                  TextFile.write(''+'\n')
            '''[:1] gets the first character of the key value and appends an * to it for the wildcard search in fnmatch
            and only take values which are not in the first row
            '''
            elif fnmatch.fnmatch(q, k[:1]+'*') and q+'1' != k:
                  TextFile.write(str(v)+'\n')
      TextFile.close()



      
