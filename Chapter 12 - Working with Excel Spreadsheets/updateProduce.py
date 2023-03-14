#! python3
# updateProduce.py - Corrects costs in the produce sales spreadsheet.

import openpyxl, os
os.chdir('C:\\Users\\segun.longe\\Documents\\Python\\Automating The Boring Stuff With Python\\Chapter 12')
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07,
                 'Celery': 1.19,
                 'Lemon': 1.27}

# Loop through the rows and update the prices
for rowNum in range(2, sheet.max_row): # skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updateProduceSales.xlsx')

